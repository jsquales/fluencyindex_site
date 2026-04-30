from __future__ import annotations

import argparse
import sys
from decimal import Decimal
from pathlib import Path
from typing import Any

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install project dependencies with "
        "`pip install -r requirements.txt` before running this script."
    ) from exc

from app.db import SessionLocal, init_db
from app.models.game24 import Game24Puzzle

# Render Shell seeding command, run from the deployed service root after the workbook
# is available there:
# python scripts/import_game24_puzzles.py "24 math game.xlsx"
DEFAULT_WORKBOOK = ROOT_DIR / "24 math game.xlsx"
SHEET_VARIANTS = {
    "Single Digits": "single_digits",
    "Integers": "integers",
}
REQUIRED_COLUMNS = {
    "style": "Style",
    "difficulty": "Difficulty",
    "n1_raw": "First",
    "n2_raw": "Second",
    "n3_raw": "Third",
    "n4_raw": "Fourth",
}
DEDUP_FIELDS = (
    "variant",
    "difficulty",
    "style",
    "n1_raw",
    "n2_raw",
    "n3_raw",
    "n4_raw",
    "source_sheet",
)


def cell_to_raw(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() or None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        decimal_value = Decimal(str(value)).normalize()
        return format(decimal_value, "f")
    return str(value).strip() or None


def header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    headers: dict[str, int] = {}
    for index, value in enumerate(header_row):
        header = cell_to_raw(value)
        if header:
            headers[header.lower()] = index
    return headers


def dedupe_key(puzzle_data: dict[str, str | None]) -> tuple[str | None, ...]:
    return tuple(puzzle_data[field] for field in DEDUP_FIELDS)


def existing_keys(db) -> set[tuple[str | None, ...]]:
    return {
        tuple(row)
        for row in db.query(*(getattr(Game24Puzzle, field) for field in DEDUP_FIELDS)).all()
    }


def import_puzzles(workbook_path: Path) -> dict[str, int]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    init_db()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    db = SessionLocal()
    seen_keys = existing_keys(db)
    summary = {
        "sheets_processed": 0,
        "rows_read": 0,
        "puzzles_inserted": 0,
        "duplicates_skipped": 0,
        "missing_required_skipped": 0,
    }

    try:
        for sheet_name, variant in SHEET_VARIANTS.items():
            if sheet_name not in workbook.sheetnames:
                continue

            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            try:
                headers = header_map(next(rows))
            except StopIteration:
                summary["sheets_processed"] += 1
                continue

            missing_columns = [
                column_name
                for column_name in REQUIRED_COLUMNS.values()
                if column_name.lower() not in headers
            ]
            if missing_columns:
                raise ValueError(
                    f"Sheet '{sheet_name}' is missing required columns: {', '.join(missing_columns)}"
                )

            summary["sheets_processed"] += 1
            for row_values in rows:
                values = {
                    field: cell_to_raw(row_values[headers[column_name.lower()]])
                    for field, column_name in REQUIRED_COLUMNS.items()
                }

                if not any(values.values()):
                    continue

                summary["rows_read"] += 1
                if not values["difficulty"] or not all(
                    values[field] for field in ("n1_raw", "n2_raw", "n3_raw", "n4_raw")
                ):
                    summary["missing_required_skipped"] += 1
                    continue

                puzzle_data = {
                    "variant": variant,
                    "difficulty": values["difficulty"],
                    "style": values["style"],
                    "n1_raw": values["n1_raw"],
                    "n2_raw": values["n2_raw"],
                    "n3_raw": values["n3_raw"],
                    "n4_raw": values["n4_raw"],
                    "source_sheet": sheet_name,
                }

                key = dedupe_key(puzzle_data)
                if key in seen_keys:
                    summary["duplicates_skipped"] += 1
                    continue

                db.add(Game24Puzzle(**puzzle_data, is_active=True))
                seen_keys.add(key)
                summary["puzzles_inserted"] += 1

        db.commit()
        return summary
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Game24 puzzles from an Excel workbook.")
    parser.add_argument(
        "workbook",
        nargs="?",
        default=str(DEFAULT_WORKBOOK),
        help=f"Path to workbook. Defaults to {DEFAULT_WORKBOOK}",
    )
    args = parser.parse_args()

    summary = import_puzzles(Path(args.workbook))
    print("Game24 puzzle import complete")
    print(f"Sheets processed: {summary['sheets_processed']}")
    print(f"Rows read: {summary['rows_read']}")
    print(f"Puzzles inserted: {summary['puzzles_inserted']}")
    print(f"Duplicates skipped: {summary['duplicates_skipped']}")
    print(f"Rows skipped due to missing required values: {summary['missing_required_skipped']}")


if __name__ == "__main__":
    main()
